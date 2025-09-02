from openpyxl import load_workbook
from database.connection import Base, engine, SessionLocal
from database.models.products_model import product_model
from database.models.product_type_model import product_type_model
from database.models.partners_model import partners_model
from database.models.material_type_model import material_type_model
from database.models.sales_history_model import sales_history_model

# файлы импорта
material_type_file = "excel/Material_type_import.xlsx"
sales_history_file = "excel/Partner_products_import.xlsx"
partners_file = "excel/Partners_import.xlsx"
product_type_file = "excel/Product_type_import.xlsx"
product_file = "excel/Products_import.xlsx"

Base.metadata.create_all(engine) # создание таблиц
session = SessionLocal() # сессия подключения к бд

#импорт типов продукции
def import_type_product(session):
    wb = load_workbook(product_type_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        type, koef = row
        if not type or not koef: # проверка на пустые значения
            continue
        session.add(product_type_model(type=type, koef=koef)) # добавление данных в таблицу
    session.commit()
    print("типы продукции добавлены")

#импорт типов материалов
def import_material_type(session):
    wb = load_workbook(material_type_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        material_type, defect = row
        session.add(material_type_model(material_type=material_type, defect=defect))
    session.commit()
    print("типы материалов добавлены")

#импорт партнеров
def import_partners(session):
    wb = load_workbook(partners_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        type, name, director, email, phone, address, inn, rating = row
        session.add(partners_model(type=type, name=name, director=director, email=email, phone=phone, address=address, inn=inn, rating=rating))
    session.commit()
    print("партнеры добавлены")

#импорт продуктов
def import_products(session):
    wb = load_workbook(product_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        fk_type, name, article, min_cost = row
        product_type = session.query(product_type_model).filter_by(type=fk_type).first() #получения типа продукции по совпаданию имен таблицы "типы продукции" и "продукция" (текущий докум)
        session.add(product_model(name=name, article=article,min_cost=min_cost,fk_type=product_type.id)) #добавление данных, где в столбец типа продукции записывается соотвествующий номер из табл "типы продукции"
    session.commit()
    print("продукты добавлены")

#импорт истории продаж
def import_sales_history(session):
    wb = load_workbook(sales_history_file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only= True):
        product_name, partner_name, count, date = row
        product = session.query(product_model).filter_by(name=product_name).first()
        partner = session.query(partners_model).filter_by(name=partner_name).first()
        session.add(sales_history_model(fk_products_id=product.id, fk_partners_id=partner.id, count=count, date=date))
    session.commit()
    print("история продаж добавлена")

#вызов импортов
def main():
    try:
        import_type_product(session)
        import_material_type(session)
        import_partners(session)
        import_products(session)
        import_sales_history(session)
    except Exception as e:
        session.rollback()
        print(f"Ошибка при импорте данных: {e}")
    finally:
        session.close()
        print("импорт завершен")

if __name__ == "__main__":
    main()
